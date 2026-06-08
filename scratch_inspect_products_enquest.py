import openpyxl

output_file = "bandeja_bom.xlsx"
wb = openpyxl.load_workbook(output_file, data_only=True)
ws = wb['Products as per Enquest']

print("All headers of Products as per Enquest:")
headers = [cell for cell in next(ws.iter_rows(values_only=True))]
print(headers)

print("\nFirst 40 rows of Products as per Enquest:")
for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row_idx == 1:
        continue
    if row_idx > 40:
        break
    if any(cell is not None for cell in row):
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        clean_dict = {k: v for k, v in row_dict.items() if v is not None}
        print(f"Row {row_idx}: {clean_dict}")
