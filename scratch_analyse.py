import urllib.request
import os
import subprocess
import sys

# Ensure requests and openpyxl are installed
try:
    import requests
except ImportError:
    print("Installing requests...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests"])
    import requests

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

url = "https://docs.google.com/spreadsheets/d/1uakHPR9BEXsEcwYrM2oXado1iS2NZKWb/export?format=xlsx"
output_file = "bandeja_bom.xlsx"

print(f"Downloading sheet from {url}...")
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.0.0 Safari/537.36'
}
response = requests.get(url, headers=headers)
print(f"Response status: {response.status_code}")
with open(output_file, "wb") as f:
    f.write(response.content)
print(f"Saved sheet to {output_file}, size: {os.path.getsize(output_file)} bytes")

# Load the workbook
wb = openpyxl.load_workbook(output_file, data_only=True)
print("Sheet names in the workbook:")
print(wb.sheetnames)

for sheet_name in wb.sheetnames:
    print(f"\n--- Sheet: {sheet_name} ---")
    ws = wb[sheet_name]
    
    # Print the first 25 rows
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > 25:
            print("... and more rows")
            break
        # Filter out completely empty rows
        if any(cell is not None for cell in row):
            print(f"Row {row_idx}: {list(row)[:12]}")
